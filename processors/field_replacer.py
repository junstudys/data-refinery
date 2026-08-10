import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from utils.path_manager import ensure_dir
from utils.text_fidelity import read_business_csv, read_business_excel


def replace_fields(
    dict_path: str = "config/dict_zh.xlsx",
    sheet_name: str = "dict",
    orig_folder: str = "mid_files/tmp_find_header_row",
    result_folder: str = "mid_files/tmp_field_replace",
    clear_output: bool = False,
    workspace_root: str | os.PathLike[str] | None = None,
) -> None:
    dict_df = pd.read_excel(dict_path, sheet_name=sheet_name)
    dict_df = dict_df.sort_values(by=["new_field", "priority"], ascending=True)

    ensure_dir(Path(result_folder), clear=clear_output, workspace_root=workspace_root)

    new_fields = dict_df["new_field"].unique()

    for filename in os.listdir(orig_folder):
        full_path = os.path.join(orig_folder, filename)
        if not os.path.isfile(full_path):
            continue

        if filename.endswith(".csv"):
            if os.path.getsize(full_path) == 0:
                print(f"Ignored empty file {filename}")
                continue
            df = read_business_csv(full_path)
            for new_field in new_fields:
                for _, row in dict_df[dict_df["new_field"] == new_field].iterrows():
                    if row["old_field"] in df.columns:
                        df.rename(
                            columns={row["old_field"]: row["new_field"]}, inplace=True
                        )
                        break
            df.to_csv(os.path.join(result_folder, filename), index=False)

        elif filename.endswith(".xlsx"):
            book = load_workbook(os.path.join(orig_folder, filename))
            writer = pd.ExcelWriter(
                os.path.join(result_folder, filename), engine="openpyxl"
            )
            writer.book = book
            for sheet in book.sheetnames:
                if book[sheet].calculate_dimension() == "A1":
                    print(f"Ignored empty sheet {sheet} in {filename}")
                    continue
                df = read_business_excel(
                    os.path.join(orig_folder, filename), sheet_name=sheet
                )
                for new_field in new_fields:
                    for _, row in dict_df[dict_df["new_field"] == new_field].iterrows():
                        if row["old_field"] in df.columns:
                            df.rename(
                                columns={row["old_field"]: row["new_field"]},
                                inplace=True,
                            )
                            break
                df.to_excel(writer, index=False, sheet_name=sheet)
            writer.save()
            writer.close()
        else:
            print(f"Ignored file {filename} as it is not a csv or xlsx file.")
