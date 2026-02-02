from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula


ERROR_VALUES = {"#N/A", "#NAME?", "#VALUE!", "#REF!", "#DIV/0!"}


@dataclass
class PreprocessConfig:
    enabled: bool = True
    in_place: bool = False
    audit_log_path: str = "logs/excel_error_audit.csv"
    error_mode: str = "keep"
    preserve_formula_text: bool = True
    preserve_if_contains_chinese: bool = True
    preserve_if_no_parentheses: bool = True


def _has_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def _has_cell_reference(text: str) -> bool:
    return bool(re.search(r"\$?[A-Z]{1,3}\$?\d+", text, re.IGNORECASE))


def _should_preserve_formula_text(formula: str, cfg: PreprocessConfig) -> bool:
    if not cfg.preserve_formula_text:
        return False
    has_parentheses = "(" in formula or ")" in formula
    has_cell_ref = _has_cell_reference(formula)
    if cfg.preserve_if_contains_chinese and _has_chinese(formula):
        return not has_parentheses and not has_cell_ref
    if cfg.preserve_if_no_parentheses and not has_parentheses:
        return not has_cell_ref
    return False


def _strip_array_formula(formula: str) -> str:
    if formula.startswith("{") and formula.endswith("}"):
        return formula[1:-1]
    return formula


def _audit_writer(path: Path) -> tuple[csv.DictWriter, object]:
    path.parent.mkdir(parents=True, exist_ok=True)
    file = path.open("w", encoding="utf-8", newline="")
    writer = csv.DictWriter(
        file,
        fieldnames=[
            "file",
            "sheet",
            "row",
            "col",
            "original_value",
            "handled_as",
            "error_type",
            "new_value",
            "timestamp",
        ],
    )
    writer.writeheader()
    return writer, file


def preprocess_excel_files(
    excel_files: Iterable[Path],
    output_folder: Path,
    cfg: PreprocessConfig,
) -> List[Path]:
    if not cfg.enabled:
        return list(excel_files)

    audit_writer, audit_file = _audit_writer(Path(cfg.audit_log_path))
    processed: List[Path] = []

    for excel_file in excel_files:
        wb = load_workbook(excel_file, data_only=False)
        changed = False
        error_mode = cfg.error_mode
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value is not None:
                        if isinstance(cell.value, ArrayFormula):
                            formula = cell.value.text or ""
                        else:
                            formula = str(cell.value)
                        if _should_preserve_formula_text(formula, cfg):
                            cleaned = _strip_array_formula(formula)
                            cell.value = cleaned
                            cell.data_type = "s"
                            cell.quotePrefix = True
                            if hasattr(sheet, "formula_attributes"):
                                sheet.formula_attributes.pop(cell.coordinate, None)
                            changed = True
                            audit_writer.writerow(
                                {
                                    "file": excel_file.name,
                                    "sheet": sheet.title,
                                    "row": cell.row,
                                    "col": cell.column,
                                    "original_value": formula,
                                    "handled_as": "formula_text",
                                    "error_type": "",
                                    "new_value": cell.value,
                                    "timestamp": datetime.now().isoformat(),
                                }
                            )
                    elif cell.data_type == "e" and cell.value in ERROR_VALUES:
                        if error_mode == "empty":
                            original = cell.value
                            cell.value = ""
                            changed = True
                            audit_writer.writerow(
                                {
                                    "file": excel_file.name,
                                    "sheet": sheet.title,
                                    "row": cell.row,
                                    "col": cell.column,
                                    "original_value": original,
                                    "handled_as": "error_value",
                                    "error_type": original,
                                    "new_value": "",
                                    "timestamp": datetime.now().isoformat(),
                                }
                            )
                            continue
                        if error_mode == "keep":
                            original = cell.value
                            cell.value = original
                            cell.data_type = "s"
                            cell.quotePrefix = True
                            changed = True
                            audit_writer.writerow(
                                {
                                    "file": excel_file.name,
                                    "sheet": sheet.title,
                                    "row": cell.row,
                                    "col": cell.column,
                                    "original_value": original,
                                    "handled_as": "error_value",
                                    "error_type": original,
                                    "new_value": cell.value,
                                    "timestamp": datetime.now().isoformat(),
                                }
                            )
                        if error_mode == "replace":
                            original = cell.value
                            cell.value = original
                            cell.data_type = "s"
                            cell.quotePrefix = True
                            changed = True
                            audit_writer.writerow(
                                {
                                    "file": excel_file.name,
                                    "sheet": sheet.title,
                                    "row": cell.row,
                                    "col": cell.column,
                                    "original_value": original,
                                    "handled_as": "error_value",
                                    "error_type": original,
                                    "new_value": cell.value,
                                    "timestamp": datetime.now().isoformat(),
                                }
                            )

        if cfg.in_place:
            target_path = excel_file
        else:
            output_folder.mkdir(parents=True, exist_ok=True)
            target_path = output_folder / excel_file.name

        if changed or cfg.in_place is False:
            wb.save(target_path)
        processed.append(target_path)

    audit_file.close()
    return processed
