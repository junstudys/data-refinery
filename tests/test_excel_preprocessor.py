from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from processors.excel_preprocessor import PreprocessConfig, preprocess_excel_files


def test_preserve_formula_text_with_chinese(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=用户1"
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="keep",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=True,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.data_type == "s"
    assert cell.value == "=用户1"


def test_preserve_formula_text_not_applied_to_real_formula(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=IF(B2<45631,1,2)"
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="keep",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=True,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.data_type == "f"


def test_error_values_are_forced_text_when_keep(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"].value = "#NAME?"
    ws["A1"].data_type = "e"
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="keep",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=False,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.data_type == "s"
    assert cell.value == "#NAME?"


def test_error_values_empty_mode(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"].value = "#DIV/0!"
    ws["A1"].data_type = "e"
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="empty",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=False,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.value in ("", None)


def test_preserve_formula_text_with_leading_dash(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=-陈成"
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="keep",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=False,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.data_type == "s"
    assert cell.value == "=-陈成"


def test_array_formula_is_converted_to_text(tmp_path: Path):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"].value = ArrayFormula("A1", "=-陈成")
    wb.save(file_path)

    processed = preprocess_excel_files(
        [file_path],
        tmp_path / "out",
        PreprocessConfig(
            enabled=True,
            in_place=False,
            audit_log_path=str(tmp_path / "audit.csv"),
            error_mode="keep",
            preserve_formula_text=True,
            preserve_if_contains_chinese=True,
            preserve_if_no_parentheses=False,
        ),
    )

    wb2 = load_workbook(processed[0], data_only=False)
    cell = wb2.active["A1"]
    assert cell.data_type == "s"
    assert cell.value == "=-陈成"
